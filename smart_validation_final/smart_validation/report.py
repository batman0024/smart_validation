from __future__ import annotations
from pathlib import Path
import pandas as pd
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

STATUS_FILLS = {
    'MISSING': 'FFC0CB', 'EMPTY': 'FFFACD', 'ACTIVE': '90EE90', 'OUTDATED':'FFDAB9',
    'NOT RUN': 'FFC0CB', 'CURRENT': '90EE90', 'OK': '90EE90', 'Warnings Found': 'ADD8E6',
    'Critical Issues Found': 'FFC0CB', 'Timing Issue': 'FFFACD', 'Run Order Issue': 'FFFACD',
    'Missing Production Program': 'FFC0CB', 'Missing Validation Program': 'FFC0CB',
    'Tool-generated (No Validation)': 'E0FFE0'
}

def autosize(ws):
    dims = {}
    for row in ws.rows:
        for cell in row:
            if cell.value is not None:
                dims[cell.column] = max(dims.get(cell.column, 0), len(str(cell.value)))
    for col, width in dims.items():
        ws.column_dimensions[get_column_letter(col)].width = min(60, width + 2)

def style_sheet_by_value(ws, column_name: str):
    header = {cell.value: idx+1 for idx, cell in enumerate(next(ws.iter_rows(min_row=1, max_row=1)))}
    if column_name not in header:
        return
    idx = header[column_name]
    for row in ws.iter_rows(min_row=2):
        cell = row[idx-1]
        val = str(cell.value) if cell.value is not None else ''
        color = STATUS_FILLS.get(val)
        if color:
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')

def write_reports(outputs: dict[str,pd.DataFrame], study, cfg) -> Path:
    docs = study.docs
    docs.mkdir(parents=True, exist_ok=True)
    from datetime import datetime
    stamp = '_' + datetime.now().strftime('%Y%m%d_%H%M%S') if cfg.get('reporting','timestamp_in_filenames', default=True) else ''
    xlsx = docs / f'Validation_Report{stamp}.xlsx'
    with pd.ExcelWriter(xlsx, engine='openpyxl') as writer:
        outputs['paths_summary'].to_excel(writer, index=False, sheet_name='Paths Summary')
        outputs['programs_tools'].to_excel(writer, index=False, sheet_name='Tools Programs')
        outputs['tnf_check'].to_excel(writer, index=False, sheet_name='TNF Check')
        outputs['all_issues'].to_excel(writer, index=False, sheet_name='Detailed Issues')
        for dom in ['SDTM','ADAM','TFL']:
            sub = outputs['final_report'].query('domain==@dom')
            if not sub.empty:
                sub.to_excel(writer, index=False, sheet_name=dom)
        ws_ps = writer.sheets.get('Paths Summary')
        if ws_ps:
            style_sheet_by_value(ws_ps, 'status')
            autosize(ws_ps)
        ws_tnf = writer.sheets.get('TNF Check')
        if ws_tnf:
            style_sheet_by_value(ws_tnf, 'status')
            autosize(ws_tnf)
        for dom in ['SDTM','ADAM','TFL']:
            ws = writer.sheets.get(dom)
            if ws:
                style_sheet_by_value(ws, 'overall_status')
                autosize(ws)
        ws_det = writer.sheets.get('Detailed Issues')
        if ws_det:
            autosize(ws_det)
    if cfg.get('reporting','write_csv', default=True):
        outputs['final_report'].to_csv(docs / f'final_report{stamp}.csv', index=False)
        outputs['all_issues'].to_csv(docs / f'issues{stamp}.csv', index=False)
    return xlsx
